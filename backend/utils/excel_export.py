import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

def generate_ai_leads_excel(leads_data):
    """
    Generates an Excel file (BytesIO object) from AI extracted leads data.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI Extracted Leads"

    headers = [
        "Name", "Phone", "Budget", "Location Preference", 
        "Property Type", "Timeline", "Interest Level", 
        "Key Requirements", "Summary", "Date"
    ]
    ws.append(headers)
    
    # Styling headers
    gold = PatternFill(start_color="C9A84C", end_color="C9A84C", fill_type="solid")
    bold = Font(bold=True, color="000000", size=11)
    for cell in ws[1]:
        cell.fill = gold
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    widths = [20, 18, 15, 20, 15, 15, 15, 30, 40, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for r in leads_data:
        ws.append([
            r.get("name", ""),
            r.get("phone", ""),
            r.get("budget", ""),
            r.get("location_preference", ""),
            r.get("property_type", ""),
            r.get("buying_timeline", ""),
            r.get("interest_level", ""),
            r.get("key_requirements", ""),
            r.get("conversation_summary", ""),
            str(r.get("created_at", ""))
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
