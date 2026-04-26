from flask import Flask, request, jsonify, Response
from flask_cors import CORS
from omnidimension import Client
import cloudinary
import cloudinary.uploader
import cloudinary.api
import requests
import json
import os
import uuid
from datetime import datetime, timedelta
from dotenv import load_dotenv
import jwt
import bcrypt
from functools import wraps
import psycopg2
from psycopg2.extras import RealDictCursor
import tempfile
import pandas as pd
from io import BytesIO

load_dotenv()

# =====================================================
#  MADHAVI ENTERPRISES - FLASK BACKEND
#  AI-Powered Real Estate Platform
# =====================================================

print('App starting...')
app = Flask(__name__)
# Enable CORS for all origins (for Vercel deployment)
CORS(app, resources={r"/*": {"origins": "*"}})

# ===== CONFIG FROM .env =====
JWT_SECRET          = os.environ.get("JWT_SECRET", "madhavi_enterprises_secret_key_2025")
OMNI_API_KEY        = os.environ.get("OMNI_API_KEY", "lqLMKAJ6b04-_xJtfWFYc_2g7MCFaOiSECW_hB6mSaonow")
OMNI_AGENT_ID       = int(os.environ.get("OMNI_AGENT_ID", "151473"))
ULTRAMSG_INSTANCE   = os.environ.get("ULTRAMSG_INSTANCE_ID", "")
ULTRAMSG_TOKEN      = os.environ.get("ULTRAMSG_TOKEN", "")
ADMIN_EMAIL         = os.environ.get("ADMIN_EMAIL", "admin@madhavienterprise.com")
ADMIN_PASSWORD      = os.environ.get("ADMIN_PASSWORD", "madhavi2025")

DATABASE_URL        = os.environ.get("DATABASE_URL")
print("DB status:", DATABASE_URL is not None)
GROQ_API_KEY        = os.environ.get("GROQ_API_KEY")

CLOUDINARY_CLOUD  = os.environ.get("CLOUDINARY_CLOUD_NAME", "djda3lldb")
CLOUDINARY_KEY    = os.environ.get("CLOUDINARY_API_KEY",    "142637855547974")
CLOUDINARY_SECRET = os.environ.get("CLOUDINARY_API_SECRET", "-5Y5ppGhCkNKwUdqnhYC4PiPz9Q")

# ===== CLOUDINARY CONFIG =====
cloudinary.config(
    cloud_name = CLOUDINARY_CLOUD,
    api_key    = CLOUDINARY_KEY,
    api_secret = CLOUDINARY_SECRET,
    secure     = True
)

# ===== OMNIDIMENSION SDK CLIENT =====
try:
    omni_client = Client(OMNI_API_KEY)
except Exception as e:
    print(f"Omnidimension Init Error: {e}")
    omni_client = None

# =====================================================
#  DATABASE - POSTGRESQL
# =====================================================
def get_pg_db():
    try:
        url = os.getenv("DATABASE_URL")
        if not url: return None
        if "?sslmode=" in url:
            return psycopg2.connect(url)
        return psycopg2.connect(url, sslmode="require")
    except Exception as e:
        print("DB Error:", e)
        return None

@app.route("/")
def home():
    return "Backend is running"


def init_pg_db():
    if not DATABASE_URL:
        print("  [!] DATABASE_URL not set. Skipping PostgreSQL init.")
        return
    try:
        conn = get_pg_db()
        c = conn.cursor(cursor_factory=RealDictCursor)
        
        # 1. Call Logs
        c.execute('''CREATE TABLE IF NOT EXISTS call_logs (
            id SERIAL PRIMARY KEY,
            customer_name VARCHAR(255),
            phone_number VARCHAR(50),
            transcription TEXT,
            summary TEXT,
            interest_type VARCHAR(50),
            property_type VARCHAR(100),
            budget VARCHAR(100),
            location VARCHAR(255),
            call_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        
        # 2. Admins
        c.execute('''CREATE TABLE IF NOT EXISTS admins (
            id SERIAL PRIMARY KEY,
            email VARCHAR(255) UNIQUE,
            password_hash TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        
        # 3. Properties
        c.execute('''CREATE TABLE IF NOT EXISTS properties (
            id SERIAL PRIMARY KEY,
            title TEXT NOT NULL,
            type TEXT DEFAULT 'Plot',
            price TEXT,
            location TEXT,
            description TEXT,
            image_urls TEXT,
            featured INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        
        # 4. Leads
        c.execute('''CREATE TABLE IF NOT EXISTS leads (
            id VARCHAR(255) PRIMARY KEY,
            property_id VARCHAR(255),
            property_title TEXT,
            name VARCHAR(255),
            phone VARCHAR(50),
            budget VARCHAR(100),
            location VARCHAR(255),
            date VARCHAR(255),
            call_status VARCHAR(50) DEFAULT 'pending',
            whatsapp_sent INTEGER DEFAULT 0
        )''')
        
        # 5. AI Extracted Leads
        c.execute('''CREATE TABLE IF NOT EXISTS ai_extracted_leads (
            id SERIAL PRIMARY KEY,
            name TEXT,
            phone TEXT,
            budget TEXT,
            location_preference TEXT,
            property_type TEXT,
            buying_timeline TEXT,
            interest_level TEXT,
            key_requirements TEXT,
            conversation_summary TEXT,
            raw_json JSONB,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        
        
        # Seed demo properties if empty
        c.execute("SELECT COUNT(*) as cnt FROM properties")
        if c.fetchone()["cnt"] == 0:
            demo = [
                ("Kammapadu Premium Plots",    "Plot",      "Rs. 15,000", "Kammapadu, Andhra Pradesh",
                 "DTCP-approved gated community plots in Kammapadu - excellent connectivity and high ROI.",
                 "https://images.unsplash.com/photo-1500382017468-9049fed747ef?w=800,https://images.unsplash.com/photo-1472214103451-9374bd1c798e?w=800", 1),
                ("Lemalle Premium Venture",    "Plot",      "Rs. 17,000", "Lemalle, Andhra Pradesh",
                 "RERA-registered venture with highway access and high appreciation potential.",
                 "https://images.unsplash.com/photo-1441974231531-c6227db76b6e?w=800,https://images.unsplash.com/photo-1433086966358-54859d0ed716?w=800", 1),
                ("Kanchincharla Exclusive",    "Plot",      "Rs. 18,000", "Kanchincharla, Andhra Pradesh",
                 "Exclusive river-view plots in Kanchincharla - limited availability, premium investment.",
                 "https://images.unsplash.com/photo-1560493676-04071c5f467b?w=800,https://images.unsplash.com/photo-1516912481808-3406841bd33c?w=800", 1),
                ("Heritage Villa",             "Villa",     "Rs. 1.8 Cr", "Vijayawada, Andhra Pradesh",
                 "Stunning 4BHK luxury villa with smart home automation and landscaped garden.",
                 "https://images.unsplash.com/photo-1600596542815-ffad4c1539a9?w=800", 0),
                ("Skyline Apartments",         "Apartment", "Rs. 85 Lakhs", "Guntur, Andhra Pradesh",
                 "Premium 3BHK apartments with rooftop club and all modern amenities.",
                 "https://images.unsplash.com/photo-1545324418-cc1a3fa10c00?w=800", 0),
                ("Riverside Individual House", "House",     "Rs. 65 Lakhs", "Krishna District, AP",
                 "Spacious individual house near Krishna river - 2400 sqft, peaceful location.",
                 "https://images.unsplash.com/photo-1568605114967-8130f3a36994?w=800", 0),
            ]
            for row in demo:
                c.execute("INSERT INTO properties (title, type, price, location, description, image_urls, featured) VALUES (%s,%s,%s,%s,%s,%s,%s)", row)
            print(f"  [+] {len(demo)} demo properties seeded")
            
        conn.commit()
        conn.close()
        print("  [+] PostgreSQL initialized for all tables")
    except Exception as e:
        print(f"  [!] PostgreSQL init failed: {e}")

# =====================================================
#  UTILITIES
# =====================================================
def normalize_encoding(text):
    if not isinstance(text, str):
        return text
    replacements = {
        "“": '"', "”": '"', "‘": "'", "’": "'", "—": "-", 
        "–": "-", "Â·": "-", "·": "-", "₹": "Rs. ",
        "\xa0": " ", "\xad": "", "â€œ": '"', "â€": '"', 
        "â€™": "'", "ðŸ": "", "ï¸": "", "âœ¦": "*"
    }
    for k, v in replacements.items():
        text = text.replace(k, v)
    cleaned_chars = [char for char in text if ord(char) < 128]
    return "".join(cleaned_chars)

def normalize_dict(d):
    if isinstance(d, dict):
        return {k: normalize_dict(v) for k, v in d.items()}
    elif isinstance(d, list):
        return [normalize_dict(v) for v in d]
    elif isinstance(d, str):
        return normalize_encoding(d)
    return d

@app.after_request
def clean_response(response):
    if response.is_json:
        try:
            data = response.get_json()
            if data:
                cleaned = normalize_dict(data)
                new_resp = jsonify(cleaned)
                new_resp.status_code = response.status_code
                for k, v in response.headers.items():
                    if k.lower() not in ('content-length', 'content-type'):
                        new_resp.headers[k] = v
                return new_resp
        except Exception:
            pass
    return response

def format_phone(phone):
    phone = str(phone).strip().replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    if phone.startswith("0"):
        phone = phone[1:]
    if not phone.startswith("+"):
        phone = "+91" + phone
    return phone

def upload_urls_to_cloudinary(raw_urls_str, folder="madhavi_properties"):
    if not raw_urls_str:
        return ""
    urls = [u.strip() for u in raw_urls_str.split(",") if u.strip()]
    result_urls = []
    for url in urls:
        try:
            if "cloudinary.com" in url or "res.cloudinary.com" in url:
                result_urls.append(url)
                continue
            r = cloudinary.uploader.upload(
                url,
                folder=folder,
                overwrite=False,
                resource_type="image"
            )
            result_urls.append(r["secure_url"])
            print(f"  [+] Cloudinary upload OK: {r['public_id']}")
        except Exception as e:
            print(f"  [!] Cloudinary upload failed for {url[:60]}: {e}")
            result_urls.append(url)
    return ",".join(result_urls)

# =====================================================
#  JWT MIDDLEWARE
# =====================================================
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth.split(" ")[1]
        if not token:
            return jsonify({"message": "Authentication required"}), 401
        try:
            jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
        except jwt.ExpiredSignatureError:
            return jsonify({"message": "Session expired, please log in again"}), 401
        except Exception as e:
            return jsonify({"message": f"Invalid token: {e}"}), 401
        return f(*args, **kwargs)
    return decorated

# =====================================================
#  AUTH ROUTES
# =====================================================
@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json() or {}
    email = data.get("email", "").strip()
    password = data.get("password", "")

    conn = get_pg_db()
    c = conn.cursor(cursor_factory=RealDictCursor)
    c.execute("SELECT * FROM admins WHERE email=%s", (email,))
    admin = c.fetchone()
    conn.close()

    if admin and bcrypt.checkpw(password.encode(), admin["password_hash"].encode()):
        token = jwt.encode({
            "admin_id": admin["id"],
            "email": admin["email"],
            "exp": datetime.utcnow() + timedelta(hours=48)
        }, JWT_SECRET, algorithm="HS256")
        return jsonify({
            "status": "success",
            "token": token,
            "agent": {"id": admin["id"], "name": "Madhavi Enterprises Admin", "email": admin["email"]}
        })
    return jsonify({"status": "error", "message": "Invalid email or password"}), 401

# =====================================================
#  PUBLIC PROPERTY ROUTES
# =====================================================
@app.route("/api/properties", methods=["GET"])
def get_properties():
    conn = get_pg_db()
    if not conn:
        return jsonify([])
    c = conn.cursor(cursor_factory=RealDictCursor)
    c.execute("SELECT * FROM properties ORDER BY featured DESC, created_at DESC")
    rows = c.fetchall()
    conn.close()
    
    for r in rows:
        imgs = r.get("image_urls", "")
        r["images"] = [i.strip() for i in imgs.split(",") if i.strip()] if imgs else []
        if 'created_at' in r and r['created_at']:
            r['created_at'] = r['created_at'].isoformat()
            
    return jsonify(rows)

@app.route("/api/properties/<int:prop_id>", methods=["GET"])
def get_property(prop_id):
    conn = get_pg_db()
    if not conn:
        return jsonify([])
    c = conn.cursor(cursor_factory=RealDictCursor)
    c.execute("SELECT * FROM properties WHERE id=%s", (prop_id,))
    row = c.fetchone()
    conn.close()
    
    if not row:
        return jsonify({"error": "Not found"}), 404
        
    imgs = row.get("image_urls", "")
    row["images"] = [i.strip() for i in imgs.split(",") if i.strip()] if imgs else []
    if 'created_at' in row and row['created_at']:
        row['created_at'] = row['created_at'].isoformat()
        
    return jsonify(row)

# =====================================================
#  ADMIN PROPERTY ROUTES
# =====================================================
@app.route("/api/admin/properties", methods=["POST"])
@token_required
def add_property():
    data = request.get_json() or {}
    data = normalize_dict(data)
    raw_urls = data.get("image_urls", "")
    cloud_urls = upload_urls_to_cloudinary(raw_urls, folder="madhavi_properties")
    conn = get_pg_db()
    c = conn.cursor()
    c.execute(
        "INSERT INTO properties (title, type, price, location, description, image_urls, featured) VALUES (%s,%s,%s,%s,%s,%s,%s)",
        (data.get("title"), data.get("type", "Plot"), data.get("price"),
         data.get("location"), data.get("description"), cloud_urls,
         1 if data.get("featured") else 0)
    )
    conn.commit()
    conn.close()
    return jsonify({"status": "success", "image_urls": cloud_urls})

@app.route("/api/admin/properties/<int:prop_id>", methods=["PUT"])
@token_required
def update_property(prop_id):
    data = request.get_json() or {}
    data = normalize_dict(data)
    raw_urls = data.get("image_urls", "")
    cloud_urls = upload_urls_to_cloudinary(raw_urls, folder="madhavi_properties")
    conn = get_pg_db()
    c = conn.cursor()
    c.execute(
        "UPDATE properties SET title=%s, type=%s, price=%s, location=%s, description=%s, image_urls=%s WHERE id=%s",
        (data.get("title"), data.get("type", "Plot"), data.get("price"),
         data.get("location"), data.get("description"), cloud_urls, prop_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"status": "success", "image_urls": cloud_urls})

@app.route("/api/admin/properties/<int:prop_id>", methods=["DELETE"])
@token_required
def delete_property(prop_id):
    conn = get_pg_db()
    if not conn:
        return jsonify([])
    c = conn.cursor()
    c.execute("DELETE FROM properties WHERE id=%s", (prop_id,))
    conn.commit()
    conn.close()
    return jsonify({"status": "success"})

# =====================================================
#  ADMIN LEADS ROUTES
# =====================================================
@app.route("/api/admin/leads", methods=["GET"])
@token_required
def get_leads():
    conn = get_pg_db()
    if not conn:
        return jsonify([])
    c = conn.cursor(cursor_factory=RealDictCursor)
    c.execute("SELECT * FROM leads ORDER BY date DESC")
    rows = c.fetchall()
    conn.close()
    return jsonify({"leads": rows})

@app.route("/api/admin/leads/export", methods=["GET"])
@token_required
def export_leads():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from flask import Response

        conn = get_pg_db()
        c = conn.cursor(cursor_factory=RealDictCursor)
        c.execute("SELECT * FROM leads ORDER BY date DESC")
        rows = c.fetchall()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Madhavi Leads"

        headers = ["Name", "Phone", "Property", "Budget", "Location", "Date", "AI Call", "WhatsApp"]
        ws.append(headers)
        gold = PatternFill(start_color="C9A84C", end_color="C9A84C", fill_type="solid")
        bold = Font(bold=True, color="000000", size=11)
        for cell in ws[1]:
            cell.fill = gold
            cell.font = bold
            cell.alignment = Alignment(horizontal="center")

        for i, w in enumerate([20, 18, 30, 18, 20, 22, 15, 15], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        for r in rows:
            ws.append([
                r.get("name",""), r.get("phone",""), r.get("property_title",""),
                r.get("budget",""), r.get("location",""), r.get("date",""),
                r.get("call_status",""), "Yes" if r.get("whatsapp_sent") else "No"
            ])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=MadhaviLeads.xlsx"}
        )
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# =====================================================
#  LEAD ENQUIRY - AI CALL + WHATSAPP
# =====================================================
@app.route("/api/enquiry", methods=["POST"])
def handle_enquiry():
    try:
        data = request.get_json() or {}
        data = normalize_dict(data)
        
        name        = data.get("name", "Friend").strip()
        raw_phone   = data.get("phone", "")
        phone       = format_phone(raw_phone)
        property_id = data.get("property_id", "")
        budget      = data.get("budget", "Not specified")
        location    = data.get("location", "Andhra Pradesh")

        property_title = data.get("property", "Property")
        price          = ""
        image1         = ""
        image2         = ""

        if property_id and str(property_id).isdigit():
            conn = get_pg_db()
            c = conn.cursor(cursor_factory=RealDictCursor)
            c.execute("SELECT * FROM properties WHERE id=%s", (int(property_id),))
            prop = c.fetchone()
            conn.close()
            if prop:
                property_title = prop["title"]
                price = prop.get("price", "")
                imgs  = [i.strip() for i in (prop.get("image_urls","")).split(",") if i.strip()]
                if imgs: image1 = imgs[0]
                if len(imgs) > 1: image2 = imgs[1]

        wa_sent = send_whatsapp(
            name, phone, property_title, price, location, image1, image2
        )

        call_status = "pending"
        try:
            call_context = {
                "name": name,
                "property": property_title,
                "price": price,
                "location": location,
                "agency": "Madhavi Enterprises"
            }
            try:
                response = omni_client.call.dispatch_call(
                    OMNI_AGENT_ID, phone, call_context=call_context
                )
                call_status = "completed"
                print(f"  [+] OmniDimension call dispatched to {phone}")
            except Exception as e1:
                print(f"  [!] dispatch_call failed: {e1}")
                call_status = "failed"

        lead_id = str(uuid.uuid4())
        conn = get_pg_db()
        c = conn.cursor()
        c.execute(
            "INSERT INTO leads (id, property_id, property_title, name, phone, budget, location, date, call_status, whatsapp_sent) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (lead_id, str(property_id), property_title, name, phone, budget, location,
             datetime.now().isoformat(), call_status, 1 if wa_sent else 0)
        )
        conn.commit()
        conn.close()

        print(f"  [+] Lead saved: {name} | {phone} | {property_title} | Call: {call_status} | WA: {wa_sent}")

        return jsonify({
            "status": "success",
            "lead_id": lead_id,
            "call_status": call_status,
            "whatsapp_sent": wa_sent
        })

    except Exception as e:
        print(f"  [!] Enquiry error: {e}")
        return jsonify({"error": str(e)}), 500

# =====================================================
#  WHATSAPP VIA ULTRAMSG
# =====================================================
def send_whatsapp(name, phone, property_title, price, location, image1="", image2=""):
    if not ULTRAMSG_INSTANCE or not ULTRAMSG_TOKEN:
        print("  [!] UltraMsg not configured")
        return False
    try:
        clean = phone.replace("whatsapp:", "").strip()
        msg = (
            f"Madhavi Enterprises\n\n"
            f"Namaste {name}! Thank you for your enquiry.\n\n"
            f"*Property:* {property_title}\n"
            f"*Price:* {price or 'Contact us'}\n"
            f"*Location:* {location}\n\n"
            f"Our AI agent is calling you now. Please pick up!\n\n"
            f"For immediate assistance:\n"
            f"Phone: +91 98765 43210\n"
            f"Web: www.madhavienterprise.com\n\n"
            f"We Don't Just Build Spaces - We Elevate Them."
        )
        url = f"https://api.ultramsg.com/{ULTRAMSG_INSTANCE}/messages/chat"
        res = requests.post(url, json={"token": ULTRAMSG_TOKEN, "to": clean, "body": msg}, timeout=10)
        result = res.json()
        if image1:
            img_url = f"https://api.ultramsg.com/{ULTRAMSG_INSTANCE}/messages/image"
            requests.post(img_url, json={"token": ULTRAMSG_TOKEN, "to": clean, "image": image1, "caption": f"{property_title} - {price}"}, timeout=10)
        if image2:
            requests.post(img_url, json={"token": ULTRAMSG_TOKEN, "to": clean, "image": image2, "caption": "Property View 2"}, timeout=10)
        sent = result.get("sent") == "true" or res.status_code == 200
        return sent
    except Exception as e:
        print(f"  [!] WhatsApp error: {e}")
        return False

# =====================================================
#  CLOUDINARY ROUTES
# =====================================================
@app.route("/api/cloudinary/images", methods=["GET"])
def get_cloudinary_images():
    folder = request.args.get("folder", "")
    max_results = min(int(request.args.get("max", 80)), 200)
    try:
        kwargs = {"resource_type": "image", "max_results": max_results, "type": "upload"}
        if folder:
            kwargs["prefix"] = folder
        result = cloudinary.api.resources(**kwargs)
        images = []
        for r in result.get("resources", []):
            images.append({
                "url":       r["secure_url"],
                "public_id": r["public_id"],
                "width":     r.get("width", 800),
                "height":    r.get("height", 600),
                "folder":    r.get("folder", ""),
                "format":    r.get("format", "jpg"),
                "created":   r.get("created_at", "")
            })
        return jsonify({"images": images, "total": len(images)})
    except Exception as e:
        print(f"  [!] Cloudinary list error: {e}")
        return jsonify({"images": [], "total": 0, "error": str(e)})

@app.route("/api/cloudinary/upload", methods=["POST"])
@token_required
def upload_image_to_cloudinary():
    data = request.get_json() or {}
    image_url = data.get("url", "").strip()
    folder    = data.get("folder", "madhavi_properties")
    if not image_url:
        return jsonify({"error": "url is required"}), 400
    try:
        if "cloudinary.com" in image_url:
            return jsonify({"secure_url": image_url, "already_hosted": True})
        r = cloudinary.uploader.upload(image_url, folder=folder, resource_type="image")
        return jsonify({"secure_url": r["secure_url"], "public_id": r["public_id"]})
    except Exception as e:
        return jsonify({"error": str(e), "secure_url": image_url}), 200

@app.route("/api/cloudinary/delete", methods=["POST"])
@token_required
def delete_cloudinary_image():
    data = request.get_json() or {}
    public_id = data.get("public_id", "")
    if not public_id:
        return jsonify({"error": "public_id required"}), 400
    try:
        result = cloudinary.uploader.destroy(public_id)
        return jsonify({"status": result.get("result", "ok")})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# =====================================================
#  AI CALL TRANSCRIPTION ROUTES
# =====================================================
@app.route("/api/admin/upload-call", methods=["POST"])
@token_required
def upload_call():
    if 'audio' not in request.files:
        return jsonify({"error": "No audio file provided"}), 400
    
    file = request.files['audio']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
        
    try:
        import os
        import json
        
        fd, temp_audio_path = tempfile.mkstemp(suffix='.mp3')
        os.close(fd)
        file.save(temp_audio_path)
        
        try:
            from groq import Groq
            client = Groq(api_key=GROQ_API_KEY)
            
            # Step 1: Transcribe using Whisper
            with open(temp_audio_path, "rb") as audio_f:
                transcription_text = client.audio.transcriptions.create(
                    file=(os.path.basename(temp_audio_path), audio_f.read()),
                    model="whisper-large-v3",
                    response_format="text"
                )
            
            # Step 2: Extract JSON using Llama 3
            prompt = """
            Analyze the following call recording and extract the key details into a JSON object.
            Return ONLY a valid JSON object with the exact following keys:
            {
                "customer_name": "extracted name or 'Unknown'",
                "phone_number": "extracted phone or 'Unknown'",
                "interest_type": "buy/rent/sell or 'Unknown'",
                "property_type": "2BHK, plot, villa, etc. or 'Unknown'",
                "budget": "extracted budget or 'Unknown'",
                "location": "extracted location or 'Unknown'",
                "summary": "1-2 sentence summary of the call"
            }
            """
            
            chat_completion = client.chat.completions.create(
                messages=[
                    {"role": "system", "content": prompt},
                    {"role": "user", "content": f"Transcript:\n{transcription_text}"}
                ],
                model="llama-3.1-8b-instant",
                response_format={"type": "json_object"}
            )
            
            extracted_data = json.loads(chat_completion.choices[0].message.content)
        finally:
            if os.path.exists(temp_audio_path):
                os.remove(temp_audio_path)
        
        conn = get_pg_db()
        c = conn.cursor()
        c.execute('''
            INSERT INTO call_logs 
            (customer_name, phone_number, transcription, summary, interest_type, property_type, budget, location)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        ''', (
            extracted_data.get('customer_name', 'Unknown'),
            extracted_data.get('phone_number', 'Unknown'),
            transcription_text,
            extracted_data.get('summary', ''),
            extracted_data.get('interest_type', 'Unknown'),
            extracted_data.get('property_type', 'Unknown'),
            extracted_data.get('budget', 'Unknown'),
            extracted_data.get('location', 'Unknown')
        ))
        log_id = c.fetchone()[0]
        conn.commit()
        conn.close()
        
        return jsonify({
            "status": "success", 
            "message": "Call processed successfully",
            "id": log_id,
            "data": extracted_data
        })
    except Exception as e:
        print(f"  [!] Upload call error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/calls", methods=["GET"])
@token_required
def get_call_logs():
    try:
        conn = get_pg_db()
        if not conn:
            return jsonify([])
        c = conn.cursor(cursor_factory=RealDictCursor)
        c.execute("SELECT id, customer_name, phone_number, summary, interest_type, property_type, budget, location, call_time FROM call_logs ORDER BY call_time DESC")
        rows = c.fetchall()
        conn.close()
        
        for r in rows:
            if 'call_time' in r and r['call_time']:
                r['call_time'] = r['call_time'].isoformat()
                
        return jsonify({"calls": rows})
    except Exception as e:
        print(f"  [!] Get calls error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/calls/<int:log_id>", methods=["GET"])
@token_required
def get_call_log_detail(log_id):
    try:
        conn = get_pg_db()
        if not conn:
            return jsonify([])
        c = conn.cursor(cursor_factory=RealDictCursor)
        c.execute("SELECT * FROM call_logs WHERE id = %s", (log_id,))
        row = c.fetchone()
        conn.close()
        
        if not row:
            return jsonify({"error": "Call log not found"}), 404
            
        if 'call_time' in row and row['call_time']:
            row['call_time'] = row['call_time'].isoformat()
            
        return jsonify({"status": "success", "data": row})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/calls/export", methods=["GET"])
@token_required
def export_call_logs():
    try:
        conn = get_pg_db()
        if not conn:
            return jsonify([])
        df = pd.read_sql_query("SELECT * FROM call_logs ORDER BY call_time DESC", conn)
        conn.close()
        
        if 'call_time' in df.columns:
            df['call_time'] = pd.to_datetime(df['call_time']).dt.tz_localize(None)
            
        buf = BytesIO()
        df.to_excel(buf, index=False, engine='openpyxl')
        buf.seek(0)
        
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=CallLogs.xlsx"}
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# =====================================================
#  ENTRY POINT
# =====================================================
if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding='utf-8', errors='replace') if hasattr(sys.stdout, 'reconfigure') else None
    print("=" * 56)
    print("  MADHAVI ENTERPRISES -- BACKEND (PRODUCTION READY)")
    print("=" * 56)
    
    # Register AI webhook and routes
    from routes.webhook import webhook_bp
    app.register_blueprint(webhook_bp)

    init_pg_db()
    print()
    print(" * Serving API with Waitress WSGI server (Port $PORT or 5001)")
    from waitress import serve
    port = int(os.environ.get("PORT", 5001))
    serve(app, host="0.0.0.0", port=port)
